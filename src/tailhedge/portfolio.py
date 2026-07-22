"""Portfolio spreadsheet in/out for the `--portfolio` sizing path.

Writes a fill-in-the-blanks .xlsx template (portfolio positions + total NAV) and
reads it back via marker-based parsing, validating that tickers aren't
duplicated and NAV covers the declared positions. Feeds `sizing.py`'s beta
regression from `advisor_cli.py` and `hedge_cli.py`.
"""

from __future__ import annotations

_TEMPLATE_EXAMPLES = [
    ("AAPL", 150_000, None, None),
    ("VOO", 330_000, None, None),
    ("SXR8", 120_000, "IBIS", "EUR"),
]


def write_template(path) -> None:
    """Writes a .xlsx template to fill in: total NAV + full-portfolio table."""
    from openpyxl import Workbook  # import lazy

    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio"
    ws["A1"] = ("Tail-hedge — fill in and save. List ALL portfolio positions "
                "(stocks, ETFs, gold, ...): the beta regression weighs each one.")
    ws["A2"] = "Cash: do NOT list it, it only counts in the total NAV below. ALL values in USD."
    ws["A3"] = ("US-listed instruments: ticker only (resolved as STK/SMART/USD). Non-US "
                "listings: fill exchange and currency (e.g. SXR8 / IBIS / EUR, "
                "VUSA / BVME.ETF / EUR).")
    ws["A4"] = "Total NAV of the portfolio (USD):"
    ws["B4"] = 1_000_000
    ws["A6"] = "ticker"
    ws["B6"] = "market_value"
    ws["C6"] = "exchange"
    ws["D6"] = "currency"
    for i, (t, v, exch, cur) in enumerate(_TEMPLATE_EXAMPLES, start=7):
        ws[f"A{i}"] = t
        ws[f"B{i}"] = v
        ws[f"C{i}"] = exch
        ws[f"D{i}"] = cur
    wb.save(path)


def load_portfolio(path) -> tuple[dict[str, float], float, dict[str, tuple[str | None, str | None]]]:
    """Reads (portfolio positions, total NAV, listings) from the Excel file, using marker-based parsing.

    Finds the column-A row that starts with 'Total NAV' → value in column B;
    finds the 'ticker' header → table (ticker, market_value, exchange, currency) below it, until the first
    empty row. Listings dict maps ticker → (exchange, currency) only for rows that declared at least one;
    values are stripped, currency is upper-cased. Raises ValueError on inconsistent input.
    """
    import zipfile

    from openpyxl import load_workbook  # import lazy
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        ws = load_workbook(path, data_only=True).active
    except (zipfile.BadZipFile, InvalidFileException):
        raise ValueError(
            f"'{path}' is not a valid .xlsx workbook. If you exported .xls or .csv, "
            "re-save it as .xlsx (Excel/LibreOffice: File > Save As > .xlsx)."
        )
    nav_total = None
    header_row = None
    for row in ws.iter_rows():
        a = row[0].value
        b = row[1].value if len(row) > 1 else None
        if isinstance(a, str):
            s = a.strip().lower()
            if s.startswith("total nav"):
                if b is None:
                    nav_total = None
                else:
                    try:
                        nav_total = float(b)
                    except (TypeError, ValueError):
                        raise ValueError(
                            f"Total NAV is not numeric ('{b}'): write it as a plain number "
                            "without thousands separators (e.g. 1000000)."
                        )
            elif s == "ticker":
                header_row = row[0].row

    positions: dict[str, float] = {}
    listings: dict[str, tuple[str | None, str | None]] = {}
    if header_row is not None:
        for row in ws.iter_rows(min_row=header_row + 1):
            t = row[0].value
            if t is None or (isinstance(t, str) and not t.strip()):
                break
            key = str(t).strip()
            if key in positions:
                raise ValueError(
                    f"Ticker '{key}' appears more than once: consolidate the positions "
                    "into a single row (duplicates are not summed, to avoid masking "
                    "copy errors)."
                )
            b = row[1].value if len(row) > 1 else None
            try:
                positions[key] = float(b)
            except (TypeError, ValueError):
                raise ValueError(f"market_value missing or not numeric for '{t}'.")
            exch = row[2].value if len(row) > 2 else None
            cur = row[3].value if len(row) > 3 else None
            exch = exch.strip() if isinstance(exch, str) and exch.strip() else None
            cur = cur.strip().upper() if isinstance(cur, str) and cur.strip() else None
            if exch or cur:
                listings[key] = (exch, cur)

    if not positions or all(v <= 0 for v in positions.values()):
        raise ValueError("At least one stock position with market_value > 0 is required.")
    if nav_total is None or nav_total <= 0:
        raise ValueError("Total NAV missing or not positive ('Total NAV ...' row).")
    if nav_total < sum(positions.values()):
        raise ValueError(
            f"Total NAV ({nav_total:,.0f}) < sum of the positions "
            f"({sum(positions.values()):,.0f}): inconsistent."
        )
    return positions, nav_total, listings
