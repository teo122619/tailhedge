import pytest
from openpyxl import Workbook

from tailhedge.portfolio import write_template, load_portfolio


def _make_xlsx(path, nav, positions, trailing_junk=False, listings=None):
    """Builds a .xlsx with the template's marker-based layout (extra rows tolerated)."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Tail-hedge — fill in and save."
    ws["A3"] = "Total NAV of the portfolio (USD):"
    ws["B3"] = nav
    ws["A5"] = "ticker"
    ws["B5"] = "market_value"
    r = 6
    for t, v in positions.items():
        ws[f"A{r}"] = t
        ws[f"B{r}"] = v
        if listings and t in listings:
            exch, cur = listings[t]
            ws[f"C{r}"] = exch
            ws[f"D{r}"] = cur
        r += 1
    if trailing_junk:
        r += 1  # empty row
        ws[f"A{r}"] = "note: this row must NOT end up among the positions"
    wb.save(path)


def test_write_template_then_load_roundtrip(tmp_path):
    path = tmp_path / "port.xlsx"
    write_template(str(path))
    positions, nav, _ = load_portfolio(str(path))
    assert nav > 0
    assert len(positions) >= 1
    assert all(v > 0 for v in positions.values())


def test_load_parses_positions_and_nav(tmp_path):
    path = tmp_path / "p.xlsx"
    _make_xlsx(path, 1_000_000, {"AAPL": 150_000, "MSFT": 120_000, "VUSA": 330_000})
    positions, nav, _ = load_portfolio(str(path))
    assert nav == 1_000_000.0
    assert positions == {"AAPL": 150_000.0, "MSFT": 120_000.0, "VUSA": 330_000.0}


def test_load_stops_at_empty_row_and_ignores_junk(tmp_path):
    path = tmp_path / "p.xlsx"
    _make_xlsx(path, 800_000, {"AAPL": 100_000}, trailing_junk=True)
    positions, nav, _ = load_portfolio(str(path))
    assert positions == {"AAPL": 100_000.0}  # the note row after the blank one is ignored


def test_load_raises_without_positions(tmp_path):
    path = tmp_path / "p.xlsx"
    _make_xlsx(path, 1_000_000, {})
    with pytest.raises(ValueError, match="position"):
        load_portfolio(str(path))


def test_load_raises_without_nav(tmp_path):
    path = tmp_path / "p.xlsx"
    _make_xlsx(path, None, {"AAPL": 100_000})
    with pytest.raises(ValueError, match="Total NAV"):
        load_portfolio(str(path))


def test_load_raises_when_nav_below_positions_sum(tmp_path):
    path = tmp_path / "p.xlsx"
    _make_xlsx(path, 50_000, {"AAPL": 100_000})  # NAV < sum
    with pytest.raises(ValueError, match="sum"):
        load_portfolio(str(path))


def test_load_raises_on_duplicate_ticker(tmp_path):
    # two lots of the same stock: NOT silently summed (it would misstate the
    # coverage of a risk-sizing tool) → error that names the ticker
    path = tmp_path / "p.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A3"] = "Total NAV of the portfolio (EUR):"
    ws["B3"] = 1_000_000
    ws["A5"] = "ticker"
    ws["B5"] = "market_value"
    ws["A6"] = "AAPL"
    ws["B6"] = 100_000
    ws["A7"] = "AAPL"
    ws["B7"] = 50_000
    wb.save(str(path))
    with pytest.raises(ValueError, match="more than once"):
        load_portfolio(str(path))


def test_load_raises_on_non_numeric_nav(tmp_path):
    # NAV written as text (e.g. "1,000,000" with separators) → named message,
    # not the raw ValueError from float()
    path = tmp_path / "p.xlsx"
    _make_xlsx(path, "one million", {"AAPL": 100_000})
    with pytest.raises(ValueError, match="Total NAV"):
        load_portfolio(str(path))


def test_load_raises_on_non_xlsx_file(tmp_path):
    # a .csv/.xls export renamed .xlsx (or any non-workbook) → clean ValueError,
    # not a raw zipfile.BadZipFile / InvalidFileException traceback.
    path = tmp_path / "not-a-workbook.xlsx"
    path.write_text("ticker,market_value\nAAPL,150000\n")
    with pytest.raises(ValueError, match="valid .xlsx"):
        load_portfolio(str(path))


def test_template_documents_and_carries_non_us_listing_columns(tmp_path):
    """The template deliberately ships a non-US example (SXR8 on IBIS2/EUR): the
    sheet must document how to fill exchange/currency for such listings, and
    the example row itself must carry that listing in columns C/D, not just
    describe it in prose."""
    import openpyxl
    from tailhedge.portfolio import load_portfolio, write_template

    path = tmp_path / "t.xlsx"
    write_template(path)
    ws = openpyxl.load_workbook(path).active
    assert ws["C6"].value == "exchange"
    assert ws["D6"].value == "currency"
    text = " ".join(str(c.value) for r in ws.iter_rows() for c in r if c.value)
    assert "exchange" in text.lower() and "currency" in text.lower()
    assert "Non-US" in text     # instructs filling exchange/currency for non-US listings

    positions, _, listings = load_portfolio(path)
    assert positions["SXR8"] == 120_000.0
    assert listings["SXR8"] == ("IBIS2", "EUR")


def test_template_has_listing_columns_and_usd_label(tmp_path):
    path = tmp_path / "port.xlsx"
    write_template(str(path))
    from openpyxl import load_workbook
    ws = load_workbook(str(path)).active
    assert "USD" in ws["A4"].value
    assert ws["C6"].value == "exchange"
    assert ws["D6"].value == "currency"


def test_load_returns_listings_for_declared_rows(tmp_path):
    path = tmp_path / "p.xlsx"
    _make_xlsx(path, 1_000_000,
               {"AAPL": 150_000, "SXR8": 120_000, "VUSA": 330_000},
               listings={"SXR8": ("ibis2", "eur"), "VUSA": (None, "EUR")})
    positions, nav, listings = load_portfolio(str(path))
    assert positions["SXR8"] == 120_000.0
    # lowercase declared exchange round-trips uppercase, symmetric with currency
    assert listings == {"SXR8": ("IBIS2", "EUR"), "VUSA": (None, "EUR")}


def test_load_old_sheet_without_listing_columns_is_unchanged(tmp_path):
    path = tmp_path / "p.xlsx"
    _make_xlsx(path, 800_000, {"AAPL": 100_000, "VOO": 300_000})
    positions, nav, listings = load_portfolio(str(path))
    assert positions == {"AAPL": 100_000.0, "VOO": 300_000.0}
    assert nav == 800_000.0
    assert listings == {}
